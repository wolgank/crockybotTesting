from distutils.sysconfig import PREFIX

from email import message

from lib2to3.pgen2 import token

from aiohttp import request

import discord

import os

import io

from PIL import Image

import requests

import openpyxl

import numpy as np

import random

import services.services

from discord import channel

from discord.ext import commands, tasks

from itertools import cycle



from discord.ext.commands.core import command



intents = discord.Intents.default()

intents.message_content = True





PREFIJO_COMANDOS="?"

MAX_NIVELES_COMANDO=50

client=commands.Bot(command_prefix=PREFIJO_COMANDOS, intents=intents)

cargaInicialBot={}

client.remove_command('help')



usuariosYaVerificados={}



@client.event

async def on_ready():

    print('Bot esta listo')

    global cargaInicialBot

    global usuariosYaVerificados

    cargaInicialBot=services.services.cargaInicial()

    usuariosYaVerificados=services.services.listarUsuariosVerificados()

    await client.change_presence(activity=discord.Activity(type=discord.ActivityType.watching,name=f'?help'))





@client.event

async def on_member_join(member):#no funciona

    for channel in client.get_all_channels():

        if channel.name=='general':

            await channel.send('Bienvenido Estudiante')



@client.event

async def on_guild_join(guild):

    for channel in guild.text_channels:

        if channel.name=="información-bot":

            await channel.send("¡Hola, soy Crocky Bot! Si tienes alguna consulta escribeme por privado con ?help para acceder al menú de consultas.\n"+

                                "Si estás usando Crocky Bot por primera vez, debes verificarte siguiendo los siguientes pasos:\n"+

                                "1) Mandandome un mensaje con ?token CODIGO, reemplazando CODIGO por tu código PUCP (Ejemplo: ?token 20173002). Al hacer esto, recibirás un correo con tu token de verificación.\n"+

                                "2) Con el token de verificación que se te fue enviado a tu correo PUCP, verificate finalmente escribiendo un mensaje con ?verify TOKEN, reemplazando TOKEN con el token de verificación (Ejemplo: ?verify 3037). Recuerda que es un código de 4 digitos.\n"+

                                "3) ¡Listo! Ya estás verificado y puedes hacerme las consultas que deseas, puedes acceder al menú con ?help. Si las opciones presentadas no responden tu duda, puedes enviar una consulta personalizada escribiéndome ?consultaPersonalizada CONSULTA, reemplazando CONSULTA con la consulta que tengas (Ejemplo: ?consultaPersonalizada Cuál es el precio del crédito?). Esto enviará tu consulta a alguien que pueda atenderla y recibirás un correo con la respuesta en los próximos días.)\n",

                                file=discord.File("screenshotsAyuda/informacionUsuariosBot.png"))

        elif channel.name=="información-bot-admin":

            pass

        elif channel.name=="bienvenida":

            pass





@client.event

async def on_message(message):

    if message.author.bot:

        return

    global cargaInicialBot

    global usuariosYaVerificados

    if not message.guild: #mensaje privado uno a uno

        if str(message.author.id) in usuariosYaVerificados:#el usuario esta verificado

            if message.content.startswith(PREFIJO_COMANDOS+"token"):

                await message.channel.send("Usted ya se encuentra verificado. Si tiene alguna duda sobre como usar los comandos escriba el comando '"+PREFIJO_COMANDOS+"help'")

            elif message.content.startswith(PREFIJO_COMANDOS+"verify"):

                await message.channel.send("Usted ya se encuentra verificado. Si tiene alguna duda sobre como usar los comandos escriba el comando '"+PREFIJO_COMANDOS+"help'")

            else:

                await comandosAlumno(message)

        else:#se debe verificar

            if message.content.startswith(PREFIJO_COMANDOS+"token"):

                await comandoEnviarToken(message)

            elif message.content.startswith(PREFIJO_COMANDOS+"verify"):

                await comandoVerificarToken(message)

            else:

                await message.channel.send("Hola, este es el bot de consultas EEGGCC. Para poder realizar sus consultas antes tiene que verificarse.\n"+

                "Para verificarse escriba el comando '"+PREFIJO_COMANDOS+"token CODIGO' cambiando la palabra 'CODIGO' por su código de estudiante. Por ejemplo:\n"

                +PREFIJO_COMANDOS+"token 20151887")

    else: #canal de servidor

        role=discord.utils.get(message.guild.roles,name="Administrador")

        if role in message.author.roles:

            await comandosAdministrador(message)



async def comandosAlumno(message):

    if message.content.startswith(PREFIJO_COMANDOS+"consultaPersonalizada"):

        message.content=" ".join(message.content.split())

        p=0

        try:

            p=message.content.index(' ')

        except:

            await message.channel.send("El comando '"+PREFIJO_COMANDOS+"consultaPersonalizada' se encuentra mal escrito. Verifica que cumpla con el formato, el cual es:\n"+

            PREFIJO_COMANDOS+"consultaPersonalizada CONSULTA donde 'CONSULTA' es la consulta que desea realizar. Por ejemplo:\n"

            +PREFIJO_COMANDOS+"consultaPersonalizada ¿El primer pago de la boleta es un pago extra a los créditos matriculados?")

            return

        consulta=message.content[p+1:]

        resultado, mensaje= services.services.consultaPersonalizada(message.author.id,consulta)

        if resultado:

            await message.channel.send(mensaje)

        else:

            await message.channel.send("No se pudo enviar su consulta debido a: "+mensaje)

    else:

        await comandosInformacion(message)

    return



async def comandosAdministrador(message):

    global cargaInicialBot

    if message.content==(PREFIJO_COMANDOS+"helpAdmin"):

        embedVar = discord.Embed(title="Ayuda para comandos de administrador", description="Aquí se encuentra la lista de los comandos disponibles solo para los administradores. (No olvidar anteponer el "+PREFIJO_COMANDOS+"):", color=0xff0000)

        embedVar.add_field(name="registrarAlumnos", value="Comando con el que se registra a los alumnos mediante un archivo .csv para que puedan ser verificados.", inline=False)

        embedVar.add_field(name="descargarFormatoRegistrarAlumnos", value="Comando con el que se descarga un archivo .csv con el formato a seguir para registrar a los alumnos", inline=False)

        embedVar.add_field(name="registrarComandos", value="Comando con el que se registra los comandos mediante un archivo .xlsx", inline=False)

        embedVar.add_field(name="descargarFormatoRegistrarComandos", value="Comando con el que se descarga un archivo .xlsx con el formato a seguir para registrar los comandos", inline=False)

        embedVar.add_field(name="reporteIndicadores",value="Comando con el que se muestra mediante un mensaje el número de alumnos verificados, número de consultas hechas, entre otras estadísticas del uso del bot.",inline=False)

        embedVar.add_field(name="reporteConsultasPersonalizadas",value="Comando con el que se descarga un archivo .csv con la lista de las consultas personalizadas hechas indicando el correo del alumno que la realizo.",inline=False)

        embedVar.add_field(name="reporteUsuarios", value="Comando con el que se descarga un archivo .csv con una lista de los alumnos registrados", inline=False)

        embedVar.add_field(name="help", value="Comando con el que se visualiza los comandos disponibles para los alumnos", inline=False)

        await message.channel.send(embed=embedVar)

    elif message.content.startswith(PREFIJO_COMANDOS+"registrarAlumnos"):

        if message.attachments==[]:

            await message.channel.send("El comando '"+PREFIJO_COMANDOS+"registrarAlumnos' debe tener adjuntado un archivo .csv con el formato dado.\n"

            +"Si no tiene el formato del archivo puede descargar un ejemplo usando el comando  '"+PREFIJO_COMANDOS+"descargarFormatoRegistrarAlumnos.\n"

            +"También puede descargar el archivo desde el campus virtual. Abajo un ejemplo de como usar el comando(el nombre del archivo no es importante):",

            file=discord.File("screenshotsAyuda/screenshotRegistrarAlumnos.png"))

        elif len(message.attachments)>1:

            await message.channel.send("El comando '"+PREFIJO_COMANDOS+"registrarAlumnos' debe tener adjuntado solo un archivo .csv con el formato dado.")

        else:

            split_v1 = str(message.attachments).split("filename='")[1]

            filename = str(split_v1).split("' ")[0]

            if filename.endswith(".csv"):

                await message.channel.send("Registrando los alumnos, por favor espere...")

                archivoBytes=await message.attachments[0].read()

                print(archivoBytes)

                archivoStr=archivoBytes.decode("utf-8")

                print(archivoStr)

                resultado, mensaje=services.services.registrarUsuariosMasivo(archivoStr)

                if resultado:

                    await message.channel.send("Registro exitoso de los alumnos.")

                else:

                    await message.channel.send("No se pudo registrar los alumnos debido a:"+mensaje)

            else:

                await message.channel.send("Error en la extensión del archivo, debe ser .csv")

    elif message.content.startswith(PREFIJO_COMANDOS+"descargarFormatoRegistrarAlumnos"):

        await message.channel.send("Este es el formato del archivo para registrar los alumnos.",file=discord.File("formatoArchivos/formatoRegistrarAlumnos.csv"))

    elif message.content.startswith(PREFIJO_COMANDOS+"registrarComandos"):

        if message.attachments==[]:

            await message.channel.send("El comando '"+PREFIJO_COMANDOS+"registrarComandos' debe tener adjuntado un archivo .xslx con el formato dado.\n"

            +"Si no tiene el formato del archivo puede descargar un ejemplo usando el comando  '"+PREFIJO_COMANDOS+"descargarForamatoRegistrarComandos.\n"

            +"Abajo un ejemplo de como usar el comando(el nombre del archivo no es importante):",

            file=discord.File("screenshotsAyuda/screenshotRegistrarAlumnos.png"))

        elif len(message.attachments)>1:

            await message.channel.send("El comando '"+PREFIJO_COMANDOS+"registrarComandos' debe tener adjuntado solo un archivo .xlsx con el formato dado.")

        else:

            split_v1 = str(message.attachments).split("filename='")[1]

            filename = str(split_v1).split("' ")[0]

            if filename.endswith(".xlsx"):

                await message.channel.send("Registrando los comandos, por favor espere...")

                archivoBytes1=await message.attachments[0].read()

                archivoXlsx=io.BytesIO(archivoBytes1)

                wb=openpyxl.load_workbook(archivoXlsx)

                resultado,mensaje=cargaMasivaComandos(wb)

                if resultado:

                    cargaInicialBot=services.services.cargaInicial()

                    await message.channel.send("Registro exitoso de los comandos.")

                else:

                    await message.channel.send("No se pudo registrar los comandos debido a:"+mensaje)

            else:

                await message.channel.send("Error en la extensión del archivo, debe ser .xlsx")

    elif message.content.startswith(PREFIJO_COMANDOS+"descargarFormatoRegistrarComandos"):

        await message.channel.send("Este es el formato del archivo para registrar los comandos.",file=discord.File("formatoArchivos/formatoRegistrarComandos.xlsx"))

    elif message.content.startswith(PREFIJO_COMANDOS+"reporteIndicadores"):

        await reporteIndicador(message)

    elif message.content.startswith(PREFIJO_COMANDOS+"reporteConsultasPersonalizadas"):

        await reporteConsultaPersonalizada(message)

    elif message.content.startswith(PREFIJO_COMANDOS+"reporteUsuarios"):

        await reporteUsuario(message)

    else:

        await comandosInformacion(message,True)

    return



async def reporteIndicador(message):

    indicadores=services.services.reporteIndicadores()

    mensajeIndicadores="Estos son las estadísticas de uso del bot:\n"

    for indicador in indicadores:

        mensajeIndicadores=mensajeIndicadores+" - "+indicador["nombre"]+": "+indicador["valor"]+"\n"

    await message.channel.send(mensajeIndicadores)



async def reporteConsultaPersonalizada(message):

    consultas=services.services.reporteConsultasPersonalizadas()

    consultasLista=[['Consulta','Fecha','Código alumno','Nombre de usuario en Discord','Correo']]

    for consulta in consultas:

        consultaLista=[consulta["consulta"],consulta["fecha"],consulta["codigo"],consulta["nombreUsuario"],consulta["correo"]]

        consultasLista.append(consultaLista)

    print(consultasLista)

    np.savetxt("listaConsultasPersonalizadas.csv",consultasLista,delimiter=";",fmt='% s')

    await message.channel.send("Este es el archivo con la lista de las consultas personalizadas realizadas hasta la fecha",file=discord.File("listaConsultasPersonalizadas.csv"))



async def reporteUsuario(message):

    usuarios=services.services.reporteUsuarios()

    usuariosLista=[['ID usuario Discord','Código alumno','Nombre de usuario en Discord','Correo','Verificado']]

    for usuario in usuarios:

        if usuario["verificado"]:

            usuarioLista=[usuario["idUsuarioPlataforma"],usuario["codigo"],usuario["nombreUsuario"],usuario["correo"],"Sí"]

        else:

            usuarioLista=["-",usuario["codigo"],"-",usuario["correo"],"No"]

        usuariosLista.append(usuarioLista)

    np.savetxt("listaUsuarios.csv",usuariosLista,delimiter=";",fmt='% s')

    await message.channel.send("Este es el archivo con la lista de los usuarios",file=discord.File("listaUsuarios.csv"))



def cargaMasivaComandos(wb):#suponiendo que el help es el comando principal y todos son subcomandos de él

    ws=wb.worksheets[0]

    dictComandos={}

    matrizAuxiliar=[[] for x in range(MAX_NIVELES_COMANDO)]



    fila, col=2,1

    dictComandos["indice"]=str(ws.cell(fila,col).value)

    dictComandos["titulo"]=ws.cell(fila,col+1).value

    dictComandos["respuesta"]=ws.cell(fila,col+3).value

    dictComandos["nombre"]=ws.cell(fila,col+2).value

    dictComandos["imagenUrl"]=ws.cell(fila,col+4).value if ws.cell(fila,col+4).value else ""



    dictComandos["subComandos"]=[]



    fila+=1

    while ws.cell(fila,col).value!=None:

        matrizAuxiliar[str(ws.cell(fila,col).value).count('.')].append({

            "indice":str(ws.cell(fila,col).value),

            "fila":fila,

        })

        fila+=1

    for primerNivelSubComandos in matrizAuxiliar[1]:#primera fila de subcomandos

        dictComandos["subComandos"].append({

            "indice":primerNivelSubComandos["indice"],

            "titulo":ws.cell(primerNivelSubComandos["fila"],col+1).value,

            "respuesta":ws.cell(primerNivelSubComandos["fila"],col+3).value,

            "nombre":ws.cell(primerNivelSubComandos["fila"],col+2).value,

            "imagenUrl":ws.cell(primerNivelSubComandos["fila"],col+4).value if ws.cell(primerNivelSubComandos["fila"],col+4).value else "",

            "subComandos":[],

        })

    print(matrizAuxiliar)

    i=2#seria el segundo nivel de subcomandos(ejemplo matricula primera opcion)

    j=0

    subComandosAux=dictComandos["subComandos"]#en este punto seria el array de los primeros subComandos

    k=0

    #subComandosAux esta tomando dictComandos de referencia, los cambios que se apliquen en este se aplicaran a dictComandos

    while matrizAuxiliar[i]!=[]:

        k=0

        while k<len(matrizAuxiliar[i]):#recorre los elementos del arreglo de la matriz(los indices)

            #recordatorio: no volver a hacerse el piola

            j=0

            while j<len(subComandosAux):#lo de adentro debe asegurar que se salga  cuando ya se halla llegado al punto de insercion

                if matrizAuxiliar[i][k]["indice"].startswith(subComandosAux[j]["indice"]):#y que se verifique el 1.1.10, 1.1.11, etc

                    if matrizAuxiliar[i][k]["indice"][len(subComandosAux[j]["indice"])] == ".":

                        subComandosAux=subComandosAux[j]["subComandos"]

                        j=0#por si acaso el nuevo subComandosAux tiene elementos, para recorrerlos de nuevo

                    else:

                        j+=1

                else:#busca en el siguiente elemento del arreglo actual de subcomandos

                    j+=1

            #en este punto subComandosAux referencia al arreglo de subCOmandos de dict donde debe ser insertado

            subComandosAux.append({

                "indice":matrizAuxiliar[i][k]["indice"],

                "titulo":ws.cell(matrizAuxiliar[i][k]["fila"],col+1).value,

                "respuesta":ws.cell(matrizAuxiliar[i][k]["fila"],col+3).value,

                "nombre":ws.cell(matrizAuxiliar[i][k]["fila"],col+2).value,

                "imagenUrl":ws.cell(matrizAuxiliar[i][k]["fila"],col+4).value if ws.cell(matrizAuxiliar[i][k]["fila"],col+4).value else "",

                "subComandos":[],

            })

            subComandosAux=dictComandos["subComandos"]#empieza la búsqueda desde cero

            k+=1

        i+=1

    dictAuxComandos=[]

    dictAuxComandos.append(dictComandos)

    resultado, mensaje=services.services.registrarComandosMasivo(dictAuxComandos)

    return resultado, mensaje



async def comandoEnviarToken(message):

    message.content=" ".join(message.content.split())

    p=0

    try:

        p=message.content.index(' ')

    except:

        await message.channel.send("El comando '"+PREFIJO_COMANDOS+"token' se encuentra mal escrito. Verifica que cumpla con el formato, el cual es:\n"+

        PREFIJO_COMANDOS+"token CODIGO donde 'CODIGO' es su código de estudiante. Por ejemplo:\n"

        +"token 20151887")

        return

    codigoEstudiante=message.content[p+1:]

    if len(codigoEstudiante)==8 and codigoEstudiante.isdigit():

        resultado,mensaje=services.services.enviarToken(message.author.id,message.author.name + "#" + message.author.discriminator,codigoEstudiante)

    else:

        await message.channel.send("El comando '"+PREFIJO_COMANDOS+"token' se encuentra mal escrito. Verifica que cumpla con el formato, el cual es:\n"+

        PREFIJO_COMANDOS+"token CODIGO donde 'CODIGO' es su codigo de estudiante. Por ejemplo:\n"

        +PREFIJO_COMANDOS+"token 20151887")

        return

    if resultado:

        await message.channel.send(mensaje+" Copia el token y utiliza el comando '"+

        PREFIJO_COMANDOS+"verify TOKEN' donde 'TOKEN' es el número enviado a tu correo. Recuerda que este número tiene 4 dígitos. Por ejemplo:\n"+

        PREFIJO_COMANDOS+"verify 9999")

    else:

        await message.channel.send("No se pudo enviar el token por el siguiente motivo: "+mensaje)

    return



async def comandoVerificarToken(message):

    global usuariosYaVerificados

    message.content=" ".join(message.content.split())

    p=0

    try:

        p=message.content.index(' ')

    except:

        await message.channel.send("El comando '"+PREFIJO_COMANDOS+"verify' se encuentra mal escrito. Verifica que cumpla con el formato, el cual es:\n"+

        PREFIJO_COMANDOS+"verify TOKEN donde 'TOKEN' es el número enviado a tu correo. Recuerda que este número tiene 4 dígitos. Por ejemplo:\n"

        +PREFIJO_COMANDOS+"verify 9999")

        return

    tokenEstudiante=message.content[p+1:]

    if len(tokenEstudiante)==4 and tokenEstudiante.isdigit():

        resultado,mensaje=services.services.verificarToken(message.author.id,message.author.name + "#" + message.author.discriminator,tokenEstudiante)

    else:

        await message.channel.send("El comando '"+PREFIJO_COMANDOS+"verify' se encuentra mal escrito. Verifica que cumpla con el formato, el cual es:\n"+

        PREFIJO_COMANDOS+"verify TOKEN donde 'TOKEN' es el número enviado a tu correo. Recuerda que este número tiene 4 dígitos. Por ejemplo:\n"

        +PREFIJO_COMANDOS+"verify 9999")

        return

    if resultado:

        usuariosYaVerificados[str(message.author.id)]=True

        await message.channel.send("Ya se encuentra verificado. Utilice el comando '"+PREFIJO_COMANDOS+"help' para que empezar a realizar sus consultas.")

    else:

        await message.channel.send("No se pudo verificar el token por el siguiente motivo: "+mensaje)

    return



async def comandosInformacion(message,esAdmin=False):

    global cargaInicialBot

    if message.content.startswith(PREFIJO_COMANDOS):

        cadenaEntrante = message.content[1:] #contiene el comando que mandó el usuario

        pos=encontrarComando(cadenaEntrante)

        if pos!=-1:#se encontro el comando

            embedVar = discord.Embed(title=cargaInicialBot[pos]["titulo"], description=cargaInicialBot[pos]["respuesta"], color=0xff0000)

            i=0

            lenSubComand = len(cargaInicialBot[pos]["subComandos"])

            while i < lenSubComand:

                embedVar.add_field(name=cargaInicialBot[pos]["subComandos"][i]["titulo"], value=cargaInicialBot[pos]["subComandos"][i]["nombre"], inline=True)

                i = i + 1

            if cargaInicialBot[pos]["imagenUrl"]!="" and cargaInicialBot[pos]["imagenUrl"]!=None:

                if cargaInicialBot[pos]["imagenUrl"].startswith("http"):

                    response=requests.get(cargaInicialBot[pos]["imagenUrl"])

                    img=io.BytesIO(response.content)

                    file=discord.File(img,"imagen.png")

                    embedVar.set_image(url="attachment://imagen.png")

                    await message.channel.send(file=file,embed=embedVar)

                else:

                    file=discord.File(cargaInicialBot[pos]["imagenUrl"],"imagen.png")

                    embedVar.set_image(url="attachment://imagen.png")

                    await message.channel.send(file=file,embed=embedVar)

            else:

                await message.channel.send(embed=embedVar)

                pass

            if not(esAdmin):

                services.services.enviarDatosUsuario(message.author.name + "#" + message.author.discriminator,cargaInicialBot[pos]["id"],message.author.id)

        else:

            if esAdmin:

                await message.channel.send("El comando no existe o se encuentra mal escrito. Si no recuerda algún comando puede escribir '"+PREFIJO_COMANDOS+"helpAdmin'.")

            else:

                await message.channel.send("El comando no existe o se encuentra mal escrito. Si no recuerda algún comando puede escribir '"+PREFIJO_COMANDOS+"help'.\n"

                +"En caso de que no encuentre lo que desea, puede escribir una consulta personalizada la cual será respondida mediante su correo electrónico de la universidad.\n"+

                "Use el comando '"+PREFIJO_COMANDOS+"consultaPersonalizada CONSULTA' donde 'CONSULTA' es la consulta que desea realizar. Por ejemplo:\n"

                +"consultaPersonalizada ¿El primer pago de la boleta es un pago extra a los créditos matriculados?")



def encontrarComando(cadena):

    global cargaInicialBot

    izq=0

    der=len(cargaInicialBot)-1

    while izq<=der:

        medio=(izq+der)//2

        if(cadena==cargaInicialBot[medio]["nombre"]):

            return medio

        elif(cadena<cargaInicialBot[medio]["nombre"]):

            der=medio-1

        else:

            izq=medio+1

    return -1





@client.event

async def on_command_error(ctx, error):

    if isinstance(error,commands.MissingRequiredArgument):#manejar errores que les falte un argumento

        await ctx.send('Faltan argumentos para el comando')

        return

    if isinstance(error,commands.CommandNotFound):#manejar error de que no existe el comando

        await ctx.send('No existe este comando.')

        return

    await ctx.send('Error')



for filename in os.listdir('./cogs'):

    if filename.endswith('.py'):

        client.load_extension(f'cogs.{filename[:-3]}')

|

client.run('ODc1MDQxNTMzNDQ1OTYzODI3.YRPv7w.thrA-Zkd-uq85Kxuj2L9AX3UZGg')
#MTE2NDY1NjExMjcwMDc2MDE2NQ.GWI8rR.FUICGDMB8t5cmya1LWPVX0RVfbqe6QHN346jbw
#ODc1MDQxNTMzNDQ1OTYzODI3.YRPv7w.thrA-Zkd-uq85Kxuj2L9AX3UZGg